from JungleScout.Jungle.Jungle_Scout import Jungle_Scout
import os
import shutil
import openpyxl
import pandas as pd


def porcess(row):
    query = row['Nom']
    chiffreAffaireVise = int(row['Chiffre Affaire visé'])
    volumeRechercheDuMotClePrincipal = ""
    profondeurDuMarche_nbrDeVendeursAvecLeCAVise = ""
    nbrDeVendeursAvec75ReviewsEtMoins = ""
    nbrDeVendeursAvecUnRatingDe4EtoilesEtMoins = ""
    nbrDeVendeursAvecDeLisingNonOptimise = ""
    nbrDeVendeursFBM = ""
    nbrDeVendeursAvecUneOffreSimilaire = int(row['le nombre de vendeur avec une offre similaire'])
    laMargeDeProfit = float(row['la marge de profit'])
    produitEstConsidereDangereuxOuToxique = row['dangereux ou toxique'].capitalize()
    produitestDansUneCategorieRestriente = row['catégorie restriente'].capitalize()
    produitDisposeDunBrevet = row["dispose d'un brevet"].capitalize()

    # on peut les ignorer
    nbDeVendeursAvec500review = ""
    BSRConstant_verifieLaSaisonnalite = ""
    constanceDesPrix_siLesPrixSontStablesOuAlaHausse = ""
    nombreDeProduitsQuiSontVendusParAmazon = ""
    verifierLeCoutParClicSurLeMotClePrincipal = ""
    nbreDOffreDeMarquePopulaire = ""
    nbrDeVendeursAvecDesVentesEnCroissances = ""
    nbrDeVendeursAvecUnTauxDeReviews5EtPlus = ""

    with Jungle_Scout(True) as bot:
        bot.Amazon_Jungle_Scout_Extension(query)

    csv_folder = os.path.join(os.getcwd(), "../CSVs")
    # Supprimer le dossier et tout son contenu
    shutil.rmtree(csv_folder)
    # Recréer le dossier vide
    os.makedirs(csv_folder)

    # Get the list of files in the download folder, sorted by modification time
    files = sorted([f for f in os.listdir(download_folder) if os.path.isfile(os.path.join(download_folder, f))],
                   key=lambda f: os.path.getmtime(os.path.join(download_folder, f)))

    # Select the last two files
    last_two_files = files[-2:]

    # Move the last two files to the current folder
    for file in last_two_files:
        source = os.path.join(download_folder, file)
        destination = os.path.join(csv_folder, file)
        shutil.move(source, destination)
        # print(f"Moved '{file}' to the '../CSVs' folder.")

    files = os.listdir(csv_folder)
    matching_files = [file for file in files if file.startswith("Keyword")]
    if matching_files:
        file_path = os.path.join(csv_folder, matching_files[-1])  # Get the path of the last matching file
        df = pd.read_csv(file_path, delimiter=',', skiprows=3)
        volumeRechercheDuMotClePrincipal = df.iloc[0, 1]
        if df.iloc[0, 7] == '---':
            verifierLeCoutParClicSurLeMotClePrincipal = ''
        else:
            verifierLeCoutParClicSurLeMotClePrincipal = float(df.iloc[0, 7].replace('€', '').replace(',', ''))
    else:
        print("Please Contact The administrator")
    matching_files = [file for file in files if file.startswith("Search Term")]
    if matching_files:
        file_path = os.path.join(csv_folder, matching_files[-1])  # Get the path of the last matching file
        df = pd.read_csv(file_path)
        # Convert the 'Évaluation' column to numeric type
        df['Évaluation'] = pd.to_numeric(df['Évaluation'], errors='coerce')
        df['Avis'] = pd.to_numeric(df['Avis'], errors='coerce')
        df['Ventes mensuelles'] = pd.to_numeric(df['Ventes mensuelles'], errors='coerce')
        df['Revenus mensuels'] = df['Revenus mensuels'].str.replace('€', '').str.replace(',', '').astype(float)

        profondeurDuMarche_nbrDeVendeursAvecLeCAVise = len(df[df['Revenus mensuels'] >= chiffreAffaireVise])
        df = df[df['Revenus mensuels'] >= chiffreAffaireVise]
        nbrDeVendeursAvec75ReviewsEtMoins = len(df[df['Avis'] <= 75])
        nbrDeVendeursAvecUnRatingDe4EtoilesEtMoins = len(df[df['Évaluation'] <= 4])
        try:
            nbrDeVendeursFBM = df["Type de vendeur"].value_counts()["FBM"]
        except:
            nbrDeVendeursFBM = 0
        nbDeVendeursAvec500review = len(df[df['Avis'] >= 500])
        nombreDeProduitsQuiSontVendusParAmazon = len(df[df['Type de vendeur'].str.upper().isin(['FBA', 'AMZ'])])
        df['Taux_reviews'] = df['Avis'] / df['Ventes mensuelles']
        filtered_df = df[df['Taux_reviews'] >= 0.05]
        nbrDeVendeursAvecUnTauxDeReviews5EtPlus = len(filtered_df)
        nbrDeVendeursAvecDeLisingNonOptimise = len(df[df['LQS'] > 4])
    else:
        print("Please Contact The administrator ")
    print("Sauvegarde Dans Le Rapport")
    file_path = "rapport.xlsx"
    workbook = openpyxl.load_workbook(file_path)
    # Get the first sheet
    first_sheet = workbook.worksheets[0]
    # Create a new sheet and copy the data from the first sheet
    sheet = workbook.copy_worksheet(first_sheet)
    sheet.title = query
    for row in first_sheet.iter_rows(values_only=True):
        sheet.append(row)
    sheet["E6"] = volumeRechercheDuMotClePrincipal
    sheet["E7"] = profondeurDuMarche_nbrDeVendeursAvecLeCAVise
    sheet["E9"] = nbrDeVendeursAvec75ReviewsEtMoins
    sheet["E10"] = nbrDeVendeursAvecUnRatingDe4EtoilesEtMoins
    sheet["E11"] = nbrDeVendeursAvecDeLisingNonOptimise
    sheet["E12"] = nbrDeVendeursFBM
    sheet["E13"] = nbrDeVendeursAvecUneOffreSimilaire
    sheet["E15"] = laMargeDeProfit
    sheet["E16"] = produitEstConsidereDangereuxOuToxique
    sheet["E17"] = produitestDansUneCategorieRestriente
    sheet["E18"] = produitDisposeDunBrevet
    sheet["E20"] = nbDeVendeursAvec500review
    sheet["E21"] = BSRConstant_verifieLaSaisonnalite
    sheet["E22"] = constanceDesPrix_siLesPrixSontStablesOuAlaHausse
    sheet["E23"] = nombreDeProduitsQuiSontVendusParAmazon
    # sheet["E24"] = verifierLeCoutParClicSurLeMotClePrincipal
    if verifierLeCoutParClicSurLeMotClePrincipal == '':
        sheet["E24"] = ""
    elif 0.01 <= verifierLeCoutParClicSurLeMotClePrincipal <= 0.50:
        sheet["E24"] = "Entre 0.01 et 0.50"
    elif 0.51 <= verifierLeCoutParClicSurLeMotClePrincipal <= 1.0:
        sheet["E24"] = "Entre 0.51 et 1.00"
    elif 1.01 <= verifierLeCoutParClicSurLeMotClePrincipal <= 1.50:
        sheet["E24"] = "Entre 1.01 et 1.50"
    elif 1.51 <= verifierLeCoutParClicSurLeMotClePrincipal <= 2.0:
        sheet["E24"] = "Entre 1.51 et 2.00"
    elif verifierLeCoutParClicSurLeMotClePrincipal >= 2.01:
        sheet["E24"] = "2.01 et +"
    else:
        sheet["E24"] = ""
    sheet["E25"] = nbreDOffreDeMarquePopulaire
    sheet["E26"] = nbrDeVendeursAvecDesVentesEnCroissances
    sheet["E27"] = nbrDeVendeursAvecUnTauxDeReviews5EtPlus

    # reformating the excel file
    if verifierLeCoutParClicSurLeMotClePrincipal != "":
        sheet["B28"] = "Opportunity Score sur 110 pts"
        value = "Avoir une note minimal de 62 pts"
        sheet.unmerge_cells("C28:E28")
        sheet["C28"] = value
        sheet.merge_cells("C28:E28")
        sheet["G28"].value = '=IF(F28="","",IF(F28<=61,CHAR(251),CHAR(74)))'
    # Save the changes
    workbook.save(file_path)
    print("Sauvegardé!")


if __name__ == '__main__':

    input_path = "input.xlsx"
    download_folder = "/home/keranis/Downloads"
    inputs = pd.read_excel(input_path)

    for index, row in inputs.iterrows():
        porcess(row)
